import io
from openpyxl.worksheet import worksheet
import Classes
from Classes import *
from openpyxl import load_workbook, Workbook
from ConsoleGreet import Greeter
from Helpers.XLMapping import *
import Config
from rich.console import Console

from Models.Job import MainJob

console = Console()


# Number 0 Fill Format
def ZFill(number):
    """ Fill A Number With A Comma"""
    try:
        return "{:,}".format(number)
    except Exception:
        return


def get_excel_sheet_as_table_new(filename):
    """
      Open Excel File With File Name && And Return Excel Table
      """
    try:
        with open(filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        workbook = load_workbook(in_mem_file, read_only=True)
        sheet = workbook.active

        return sheet

    except FileNotFoundError:
        return False


def create_xl_file(file_name, **kwarg):
    """
    If There Is No File In Desktop,
    Create One
    """
    filename = file_name
    workbook = Workbook()
    sheet = workbook.active
    sheet.sheet_view.rightToLeft = True

    for cell, value in kwarg.items():
        sheet[f"{cell}"] = value

    workbook.save(filename=filename)


def populate_jobs_from_xlsheet(sheet):
    """
    Create MainJob Objects From Excel Sheet
    """
    jobs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # If Empty Row
        if row[0] is None or row[2] is None or row[6] is None:
            continue
        job = MainJob(row[JOB_DATE], row[JOB_DRIVER], row[JOB_CX], row[JOB_GIDUL], row[JOB_ZAN], row[JOB_PLANTS],
                      row[JOB_MAGASH], row[JOB_PASSPORT], row[JOB_AVG])
        jobs.append(job)
    return jobs


def write_jobs_to_bartender_file(filename, jobs, is_job_return=False):
    """
    Write Jobs To Excel File - For Each Job,
    Write Up To 9 Passports Per Job Row In Bartender Excel
    """
    workbook = load_workbook(filename)
    sheet = workbook.active
    counter = 2

    for job in jobs:
        # Sheet Columns Up To 9 Passports
        columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W'
            , 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM'
            , 'AN', 'AO', 'AP']

        num_of_passports = len(job.list_of_passports_objects)

        # Even If No Passports Need To Write This ToSheet
        sheet[f"A{counter}"] = f"{job.job_gidul} {job.job_zan} - {ZFill(job.job_plants)}"
        sheet[f"B{counter}"] = f"{job.job_cx} {job.job_magash} {Config.Heb_magash}"
        sheet[f"C{counter}"] = f"{job.job_date}"
        sheet[f"D{counter}"] = f"{job.num_of_bartender_stickers}"
        sheet[f"AR{counter}"] = f"{job.job_gidul}"
        sheet[f"AS{counter}"] = f"{job.job_zan}"
        if is_job_return:
            job.is_job_need_to_return = True
        sheet[f"AT{counter}"] = f"{job.is_job_need_to_return}"
        sheet[f"AU{counter}"] = f"{job.id}"

        # If No Passports
        if num_of_passports == 0:
            continue
        # If More Than One Passport
        else:
            # Every Passport Needs 4 Columns
            total_columns_per_passport = num_of_passports * 4
            # Slice Columns For Passports(Each Passport - 4 Columns)
            columns = columns[:total_columns_per_passport]
            # Counter To Move 1 Column (4 for 1 passports)
            # To Iterate In Passport Obj List
            passport_counter = 0
            # Passport Property
            column = 0
            while column < len(columns):
                # For Every Passport In Job
                try:
                    sheet[
                        f"{columns[column]}{counter}"] = f"{job.list_of_passports_objects[passport_counter].passport_num} "
                    column = column + 1
                    sheet[f"{columns[column]}{counter}"] = f"{job.list_of_passports_objects[passport_counter].hamama}"
                    column = column + 1
                    sheet[f"{columns[column]}{counter}"] = f"{job.list_of_passports_objects[passport_counter].gamlon}"
                    column = column + 1
                    sheet[
                        f"{columns[column]}{counter}"] = f"{job.list_of_passports_objects[passport_counter].job_num_of_magash}"
                    column = column + 1

                    # Move tTo Next Passport
                    passport_counter += 1

                except IndexError:
                    continue
        if job.is_page2:
            sheet[f"AO{counter}"] = f"{job.is_page2}"
        if job.is_page3:
            sheet[f"AP{counter}"] = f"{job.is_page3}"
        if job.is_small_cage:
            sheet[f"AQ{counter}"] = f"{job.is_small_cage}"
        counter += 1

    workbook.save(filename=filename)


def clean_sheet(filename):
    """
    Clean Sheet Before Writing Values
    """
    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    workbook = load_workbook(in_mem_file)
    sheet = workbook.active
    for row in sheet['A2:AU500']:
        for cell in row:
            cell.value = None
    # for row in sheet.iter_rows(min_row=2, values_only=True):
    #     row = None

    workbook.save(Config.Files.EXCEL_TO_BARTENDER)


def return_job(filename, jobs):
    pass


# deprecated
def write_jobs_for_bartender_Excel(filename, ListMainJob):
    """
    Write Jobs To Excel File
    """
    workbook = load_workbook(filename)
    sheet = workbook.active
    counter = 2
    for main_job in ListMainJob:
        sheet[f"A{counter}"] = f"{main_job.job_gidul} {main_job.job_zan} - {ZFill(main_job.job_plants)}"
        sheet[f"B{counter}"] = f"{main_job.job_cx} {main_job.job_magash}"
        sheet[f"C{counter}"] = f"{main_job.job_date}"
        sheet[f"D{counter}"] = f"{main_job.num_of_bartender_stickers}"
        # If No Passports In Job - Continue Loop
        if len(main_job.list_of_passports_objects) == 0:
            continue
        # In Each Job Write Passport To Excel
        for passport in main_job.list_of_passports_objects:
            # If We Have 1-6 Passports
            if main_job.is_page2:
                sheet[f"C{counter}"] = f"{main_job.job_date}"
                sheet[f"C{counter}"] = f"{main_job.job_date}"
                sheet[f"C{counter}"] = f"{main_job.job_date}"
                sheet[f"C{counter}"] = f"{main_job.job_date}"
                pass
        counter += 1

    workbook.save(filename=filename)


# deprecated
def get_data_from_excel_file_Current(filename):
    """
       Open Excel File With File Name, Start Col, End Col,
       Start Row, End Row &&  Return List Of Job Objects
       To Print
       """
    try:
        workbook = load_workbook(filename, read_only=True)
        sheet = workbook.active

        jobs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            job = MainJob(row[JOB_DATE], row[JOB_DRIVER], row[JOB_CX], row[JOB_GIDUL], row[JOB_ZAN], row[JOB_PLANTS],
                          row[JOB_MAGASH], row[JOB_PASSPORT], row[JOB_AVG])
            if job.job_date is None or job.job_cx is None or job.job_magash is None:
                continue
            jobs.append(job)
        return jobs
    except FileNotFoundError:
        console.print(Greeter.rev(Config.Heb_No_Xl) + "\n", style=Config.Styles.RedOnWhite)
        console.print(Greeter.rev(Config.Heb_Make_New_Xl) + "\n", style=Config.Styles.WhiteOnBlue)
        create_xl_file(Config.Files.FILE_TO_CREATE, A1="job_date", B1="driver", C1="job_cx", D1="job_gidul",
                       E1="job_zan", F1="job_plants"
                       , G1="job_magash", H1="job_passport", I1="job_avg")

        console.print(Greeter.rev(Config.Heb_Ok) + "\n", style=Config.Styles.GreenOnWhite)
        get_data_from_excel_file_Current(Config.Files.FILE_TO_CREATE)


# deprecated
def get_excel_sheet_as_table(filename=None, sheet_name=None, is_create=None):
    """
    Open Excel File With File Name, Start Col, End Col,
    Start Row, End Row &&  Return List Of Job Objects
    To Print
    """
    try:
        if filename is None:
            return print("[-] File Name Was Not Given!")
        workbook = load_workbook(filename, read_only=True)
        if sheet_name is None or True:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
        return sheet
    except FileNotFoundError:
        if is_create is not None:
            print("[!] File Not Found!")
        console.print(Greeter.rev(Config.Heb_No_Xl) + "\n", style=Config.Styles.RedOnWhite)
        console.print(Greeter.rev(Config.Heb_Make_New_Xl) + "\n", style=Config.Styles.WhiteOnBlue)
        create_xl_file(Config.Files.FILE_TO_CREATE, A1="job_date", B1="driver", C1="job_cx", D1="job_gidul",
                       E1="job_zan", F1="job_plants"
                       , G1="job_magash", H1="job_passport", I1="job_avg")

        console.print(Greeter.rev(Config.Heb_Ok) + "\n", style=Config.Styles.GreenOnWhite)

        return False
